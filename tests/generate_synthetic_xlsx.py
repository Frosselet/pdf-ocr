#!/usr/bin/env python3
"""Generate synthetic XLSX test files in inputs/xlsx/synthetic/.

These files are used by tests/test_xlsx_extractor.py.
Run:  uv run python tests/generate_synthetic_xlsx.py

Personas:
  P1 "The Tidy Analyst" — clean single-table sheets, mixed types
  P2 "The Merger"       — hierarchical merged headers
  P3 "The Multi-Tasker" — multiple tables per sheet
  P4 "The Formatter"    — visual styling, hidden content, date formats
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "inputs" / "xlsx" / "synthetic"


def _save(wb: Workbook, name: str) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    p = OUT / f"{name}.xlsx"
    wb.save(p)
    print(f"  {p.name}")


# ===================================================================
# P1: "The Tidy Analyst" — clean baseline
# ===================================================================


def build_p1_simple_financial() -> None:
    """1 sheet, 5 cols, 10 rows, no merges."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"

    headers = ["Company", "Revenue", "Profit", "Employees", "Founded"]
    ws.append(headers)

    data = [
        ["Acme Corp", 50000, 12000, 150, 1998],
        ["Beta Ltd", 32000, 8500, 85, 2005],
        ["Gamma Inc", 78000, 21000, 320, 1990],
        ["Delta SA", 15000, 3200, 42, 2012],
        ["Epsilon GmbH", 92000, 28000, 510, 1985],
        ["Zeta LLC", 41000, 9800, 120, 2001],
        ["Eta Pty", 63000, 17500, 230, 1995],
        ["Theta AG", 28000, 6100, 68, 2010],
        ["Iota SpA", 55000, 14200, 180, 1992],
        ["Kappa AB", 37000, 8900, 95, 2008],
    ]
    for row in data:
        ws.append(row)

    _save(wb, "p1_simple_financial")


def build_p1_inventory_mixed() -> None:
    """7 cols with mixed types (text/number/date/enum)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["Item", "Category", "Quantity", "Unit Price", "Last Restocked", "Status", "Warehouse"]
    ws.append(headers)

    data = [
        ["Widget A", "Hardware", 500, 12.50, date(2025, 11, 15), "In Stock", "NYC"],
        ["Widget B", "Hardware", 0, 8.75, date(2025, 9, 3), "Out of Stock", "LA"],
        ["Gadget X", "Electronics", 1200, 45.00, date(2026, 1, 20), "In Stock", "NYC"],
        ["Part 7Z", "Components", 8500, 0.35, date(2025, 12, 1), "In Stock", "Chicago"],
        ["Module Q", "Electronics", 75, 120.00, date(2025, 10, 8), "Low Stock", "LA"],
        ["Bolt M8", "Hardware", 25000, 0.05, date(2026, 2, 14), "In Stock", "Chicago"],
        ["Sensor K", "Electronics", 300, 28.90, date(2025, 8, 22), "In Stock", "NYC"],
        ["Cable 5m", "Accessories", 4200, 3.15, date(2026, 1, 5), "In Stock", "LA"],
        ["Cover XL", "Accessories", 180, 7.60, date(2025, 7, 30), "Low Stock", "Chicago"],
        ["Relay R3", "Components", 960, 15.40, date(2025, 11, 28), "In Stock", "NYC"],
    ]
    for row in data:
        ws.append(row)

    # Format date column
    for row_num in range(2, 12):
        ws.cell(row=row_num, column=5).number_format = "YYYY-MM-DD"

    _save(wb, "p1_inventory_mixed")


def build_p1_time_series() -> None:
    """Date column + 4 numeric columns, 30 rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TimeSeries"

    headers = ["Date", "Temperature", "Humidity", "Pressure", "Wind Speed"]
    ws.append(headers)

    import random
    random.seed(42)
    base_date = date(2025, 1, 1)
    for i in range(30):
        d = date(2025, 1 + i // 30, 1 + i % 28)
        ws.append([
            d,
            round(15 + random.gauss(0, 5), 1),
            round(60 + random.gauss(0, 10), 1),
            round(1013 + random.gauss(0, 5), 1),
            round(12 + random.gauss(0, 3), 1),
        ])
        ws.cell(row=i + 2, column=1).number_format = "YYYY-MM-DD"

    _save(wb, "p1_time_series")


def build_p1_multi_sheet() -> None:
    """3 sheets with different structures."""
    wb = Workbook()

    # Sheet 1: Sales
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Product", "Region", "Amount"])
    for product, region, amount in [
        ("Alpha", "North", 1500), ("Alpha", "South", 1200),
        ("Beta", "North", 800), ("Beta", "South", 950),
        ("Gamma", "North", 2100), ("Gamma", "South", 1800),
    ]:
        ws1.append([product, region, amount])

    # Sheet 2: Targets
    ws2 = wb.create_sheet("Targets")
    ws2.append(["Quarter", "Target", "Actual", "Variance"])
    for q, t, a in [("Q1", 5000, 4800), ("Q2", 5500, 5700), ("Q3", 6000, 5900), ("Q4", 7000, 7200)]:
        ws2.append([q, t, a, a - t])

    # Sheet 3: Regions
    ws3 = wb.create_sheet("Regions")
    ws3.append(["Code", "Name", "Manager", "Headcount", "Budget"])
    for code, name, mgr, hc, budget in [
        ("N", "North", "Alice", 45, 500000),
        ("S", "South", "Bob", 38, 420000),
        ("E", "East", "Carol", 52, 610000),
        ("W", "West", "Dave", 30, 350000),
    ]:
        ws3.append([code, name, mgr, hc, budget])

    _save(wb, "p1_multi_sheet")


# ===================================================================
# P2: "The Merger" — hierarchical merged headers
# ===================================================================


def build_p2_budget_merged() -> None:
    """2-row header: 'Revenue'/'Cost' span 2 cols each."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"

    # Row 1: merged headers
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    ws["A1"] = "Department"
    ws["B1"] = "Revenue"
    ws["D1"] = "Cost"

    # Row 2: sub-headers
    ws["A2"] = "Department"
    ws["B2"] = "Q1"
    ws["C2"] = "Q2"
    ws["D2"] = "Q1"
    ws["E2"] = "Q2"

    # Data rows
    data = [
        ["Engineering", 450, 520, 380, 410],
        ["Marketing", 300, 350, 250, 280],
        ["Sales", 600, 680, 200, 220],
        ["Operations", 150, 170, 400, 430],
        ["HR", 80, 90, 120, 135],
    ]
    for ri, row in enumerate(data, start=3):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    _save(wb, "p2_budget_merged")


def build_p2_deep_hierarchy() -> None:
    """3-row header: Year > Quarter > Month."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DeepHeaders"

    # Row 1: Year spans
    ws.merge_cells("B1:G1")
    ws.merge_cells("H1:M1")
    ws["A1"] = "Region"
    ws["B1"] = "2025"
    ws["H1"] = "2026"

    # Row 2: Quarter spans
    ws.merge_cells("B2:D2")
    ws.merge_cells("E2:G2")
    ws.merge_cells("H2:J2")
    ws.merge_cells("K2:M2")
    ws["A2"] = "Region"
    ws["B2"] = "Q1"
    ws["E2"] = "Q2"
    ws["H2"] = "Q1"
    ws["K2"] = "Q2"

    # Row 3: Month labels
    ws["A3"] = "Region"
    months_2025_q1 = ["Jan", "Feb", "Mar"]
    months_2025_q2 = ["Apr", "May", "Jun"]
    months_2026_q1 = ["Jan", "Feb", "Mar"]
    months_2026_q2 = ["Apr", "May", "Jun"]
    for i, m in enumerate(months_2025_q1):
        ws.cell(row=3, column=2 + i, value=m)
    for i, m in enumerate(months_2025_q2):
        ws.cell(row=3, column=5 + i, value=m)
    for i, m in enumerate(months_2026_q1):
        ws.cell(row=3, column=8 + i, value=m)
    for i, m in enumerate(months_2026_q2):
        ws.cell(row=3, column=11 + i, value=m)

    # Data rows
    import random
    random.seed(99)
    regions = ["North", "South", "East", "West", "Central"]
    for region in regions:
        row_data = [region] + [random.randint(100, 500) for _ in range(12)]
        ws.append(row_data)

    _save(wb, "p2_deep_hierarchy")


def build_p2_cross_tab() -> None:
    """Merged row AND column labels."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CrossTab"

    # Row merges (column headers)
    ws.merge_cells("C1:D1")
    ws.merge_cells("E1:F1")
    ws["A1"] = ""
    ws["B1"] = ""
    ws["C1"] = "Product A"
    ws["E1"] = "Product B"

    ws["A2"] = ""
    ws["B2"] = ""
    ws["C2"] = "Revenue"
    ws["D2"] = "Cost"
    ws["E2"] = "Revenue"
    ws["F2"] = "Cost"

    # Column merges (row headers)
    ws.merge_cells("A3:A4")
    ws.merge_cells("A5:A6")
    ws["A3"] = "North"
    ws["A5"] = "South"
    ws["B3"] = "Urban"
    ws["B4"] = "Rural"
    ws["B5"] = "Urban"
    ws["B6"] = "Rural"

    # Data
    data = [
        [500, 300, 400, 250],
        [200, 120, 150, 90],
        [450, 280, 350, 200],
        [180, 100, 130, 70],
    ]
    for ri, row in enumerate(data, start=3):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 3, value=val)

    _save(wb, "p2_cross_tab")


def build_p2_irregular_merges() -> None:
    """Non-rectangular merges, single-cell holes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Irregular"

    # Header row 1: one merge plus a gap
    ws.merge_cells("A1:B1")
    ws["A1"] = "Category"
    # C1 intentionally empty (gap)
    ws.merge_cells("D1:F1")
    ws["D1"] = "Metrics"

    # Header row 2: sub-labels
    ws["A2"] = "Group"
    ws["B2"] = "Type"
    ws["C2"] = "Code"
    ws["D2"] = "Value A"
    ws["E2"] = "Value B"
    ws["F2"] = "Value C"

    # Data rows
    data = [
        ["Alpha", "X", "A1", 100, 200, 300],
        ["Alpha", "Y", "A2", 110, 210, 310],
        ["Beta", "X", "B1", 150, 250, 350],
        ["Beta", "Z", "B2", 160, 260, 360],
    ]
    for row in data:
        ws.append(row)

    _save(wb, "p2_irregular_merges")


# ===================================================================
# P3: "The Multi-Tasker" — multiple tables per sheet
# ===================================================================


def build_p3_two_tables_blank_rows() -> None:
    """2 tables separated by 3 blank rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TwoTables"

    # Table 1: rows 1-6
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["C1"] = "Grade"
    data1 = [
        ["Alice", 92, "A"],
        ["Bob", 78, "B"],
        ["Carol", 85, "B+"],
        ["Dave", 96, "A+"],
        ["Eve", 71, "C+"],
    ]
    for ri, row in enumerate(data1, start=2):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    # 3 blank rows (7, 8, 9)

    # Table 2: rows 10-14
    ws["A10"] = "City"
    ws["B10"] = "Population"
    ws["C10"] = "Area"
    ws["D10"] = "Density"
    data2 = [
        ["London", 8982000, 1572, 5713],
        ["Paris", 2161000, 105, 20580],
        ["Berlin", 3645000, 892, 4088],
        ["Madrid", 3223000, 604, 5335],
    ]
    for ri, row in enumerate(data2, start=11):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    _save(wb, "p3_two_tables_blank_rows")


def build_p3_tables_with_titles() -> None:
    """Title rows above each table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TitledTables"

    # Title 1 (row 1)
    ws["A1"] = "Employee Performance"
    ws["A1"].font = Font(bold=True, size=14)

    # Table 1 (rows 2-6)
    for ci, h in enumerate(["Name", "Department", "Rating"], start=1):
        ws.cell(row=2, column=ci, value=h)
    data1 = [
        ["Alice", "Engineering", 4.5],
        ["Bob", "Marketing", 3.8],
        ["Carol", "Finance", 4.2],
        ["Dave", "Engineering", 4.7],
    ]
    for ri, row in enumerate(data1, start=3):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    # 2 blank rows (7, 8)

    # Title 2 (row 9)
    ws["A9"] = "Department Summary"
    ws["A9"].font = Font(bold=True, size=14)

    # Table 2 (rows 10-13)
    for ci, h in enumerate(["Department", "Headcount", "Avg Rating"], start=1):
        ws.cell(row=10, column=ci, value=h)
    data2 = [
        ["Engineering", 25, 4.3],
        ["Marketing", 12, 3.9],
        ["Finance", 8, 4.1],
    ]
    for ri, row in enumerate(data2, start=11):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    _save(wb, "p3_tables_with_titles")


def build_p3_tables_with_footnotes() -> None:
    """Table, footnotes, blank rows, table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Footnotes"

    # Table 1 (rows 1-5)
    for ci, h in enumerate(["Item", "Value", "Note"], start=1):
        ws.cell(row=1, column=ci, value=h)
    data1 = [
        ["Revenue", 50000, "Audited"],
        ["Expenses", 38000, "Estimated*"],
        ["Profit", 12000, ""],
        ["Tax", 3600, "30% rate"],
    ]
    for ri, row in enumerate(data1, start=2):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    # Footnotes (rows 6-7)
    ws["A6"] = "* Subject to revision"
    ws["A7"] = "Source: Annual Report 2025"

    # 2 blank rows (8, 9)

    # Table 2 (rows 10-14)
    for ci, h in enumerate(["Category", "Budget", "Actual", "Variance"], start=1):
        ws.cell(row=10, column=ci, value=h)
    data2 = [
        ["Marketing", 15000, 14200, -800],
        ["R&D", 20000, 22000, 2000],
        ["Operations", 10000, 9500, -500],
        ["Admin", 5000, 4800, -200],
    ]
    for ri, row in enumerate(data2, start=11):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    _save(wb, "p3_tables_with_footnotes")


def build_p3_side_by_side() -> None:
    """2 tables side by side, 2 blank columns between."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SideBySide"

    # Left table (cols A-C, rows 1-5)
    ws["A1"] = "Product"
    ws["B1"] = "Sales"
    ws["C1"] = "Returns"
    left_data = [
        ["Widget", 1200, 45],
        ["Gadget", 800, 20],
        ["Doohickey", 350, 10],
        ["Thingamajig", 950, 30],
    ]
    for ri, row in enumerate(left_data, start=2):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 1, value=val)

    # 2 blank columns (D, E)

    # Right table (cols F-H, rows 1-5)
    ws["F1"] = "Region"
    ws["G1"] = "Revenue"
    ws["H1"] = "Growth"
    right_data = [
        ["North", 50000, 0.12],
        ["South", 42000, 0.08],
        ["East", 38000, 0.15],
        ["West", 55000, 0.10],
    ]
    for ri, row in enumerate(right_data, start=2):
        for ci, val in enumerate(row):
            ws.cell(row=ri, column=ci + 6, value=val)

    _save(wb, "p3_side_by_side")


# ===================================================================
# P4: "The Formatter" — visual styling, hidden content
# ===================================================================


def build_p4_kpi_dashboard() -> None:
    """Conditional formatting with red/yellow/green fills."""
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"

    headers = ["KPI", "Target", "Actual", "Status"]
    ws.append(headers)
    # Bold headers with fill
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ci in range(1, 5):
        cell = ws.cell(row=1, column=ci)
        cell.fill = header_fill
        cell.font = header_font

    kpis = [
        ("Revenue Growth", 0.15, 0.18, "green"),
        ("Customer Satisfaction", 4.5, 4.2, "yellow"),
        ("Employee Retention", 0.90, 0.85, "red"),
        ("Cost Reduction", 0.10, 0.12, "green"),
        ("Market Share", 0.25, 0.22, "yellow"),
        ("Quality Score", 95, 97, "green"),
    ]

    fills = {
        "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for ri, (kpi, target, actual, status) in enumerate(kpis, start=2):
        ws.cell(row=ri, column=1, value=kpi)
        ws.cell(row=ri, column=2, value=target)
        ws.cell(row=ri, column=3, value=actual)
        status_cell = ws.cell(row=ri, column=4, value=status.upper())
        status_cell.fill = fills[status]

    _save(wb, "p4_kpi_dashboard")


def build_p4_banded_report() -> None:
    """Zebra rows, bold headers, bordered grid."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["ID", "Description", "Category", "Amount", "Date"]
    ws.append(headers)

    # Bold headers with bottom border
    thick_side = Side(style="thick")
    thin_side = Side(style="thin")
    for ci in range(1, 6):
        cell = ws.cell(row=1, column=ci)
        cell.font = Font(bold=True, size=12)
        cell.border = Border(bottom=thick_side)

    data = [
        [1001, "Office Supplies", "Admin", 250.00, date(2025, 1, 15)],
        [1002, "Server Hosting", "IT", 1500.00, date(2025, 1, 20)],
        [1003, "Training Course", "HR", 800.00, date(2025, 2, 5)],
        [1004, "Marketing Campaign", "Sales", 3200.00, date(2025, 2, 12)],
        [1005, "Software License", "IT", 450.00, date(2025, 3, 1)],
        [1006, "Travel Expenses", "Sales", 1100.00, date(2025, 3, 15)],
        [1007, "Equipment Repair", "Operations", 680.00, date(2025, 4, 2)],
        [1008, "Consulting Fees", "Finance", 2500.00, date(2025, 4, 18)],
    ]

    band_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for ri, row_data in enumerate(data, start=2):
        for ci, val in enumerate(row_data):
            cell = ws.cell(row=ri, column=ci + 1, value=val)
            cell.border = Border(
                top=thin_side, bottom=thin_side,
                left=thin_side, right=thin_side,
            )
            if ri % 2 == 0:
                cell.fill = band_fill
        ws.cell(row=ri, column=5).number_format = "YYYY-MM-DD"

    _save(wb, "p4_banded_report")


def build_p4_hidden_columns() -> None:
    """2 hidden columns with formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "HiddenCols"

    headers = ["Product", "Base Price", "Tax Rate", "Tax Amount", "Final Price"]
    ws.append(headers)

    data = [
        ["Widget A", 100, 0.10],
        ["Widget B", 250, 0.15],
        ["Gadget X", 500, 0.20],
        ["Part Y", 75, 0.10],
        ["Module Z", 1200, 0.15],
    ]

    for ri, (product, base, rate) in enumerate(data, start=2):
        ws.cell(row=ri, column=1, value=product)
        ws.cell(row=ri, column=2, value=base)
        ws.cell(row=ri, column=3, value=rate)
        ws.cell(row=ri, column=4, value=base * rate)  # Tax Amount
        ws.cell(row=ri, column=5, value=base * (1 + rate))  # Final Price

    # Hide columns C (Tax Rate) and D (Tax Amount)
    ws.column_dimensions["C"].hidden = True
    ws.column_dimensions["D"].hidden = True

    _save(wb, "p4_hidden_columns")


def build_p4_date_formats() -> None:
    """Same date in 4 regional formats."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DateFormats"

    headers = ["Event", "ISO Date", "US Date", "EU Date", "Custom Date"]
    ws.append(headers)

    events = [
        ("Launch", datetime(2025, 3, 15)),
        ("Review", datetime(2025, 6, 30)),
        ("Deadline", datetime(2025, 12, 25)),
        ("Renewal", datetime(2026, 1, 1)),
        ("Audit", datetime(2025, 9, 10)),
    ]

    for ri, (event, dt) in enumerate(events, start=2):
        ws.cell(row=ri, column=1, value=event)
        # ISO
        c_iso = ws.cell(row=ri, column=2, value=dt)
        c_iso.number_format = "YYYY-MM-DD"
        # US
        c_us = ws.cell(row=ri, column=3, value=dt)
        c_us.number_format = "MM/DD/YYYY"
        # EU
        c_eu = ws.cell(row=ri, column=4, value=dt)
        c_eu.number_format = "DD.MM.YYYY"
        # Custom
        c_custom = ws.cell(row=ri, column=5, value=dt)
        c_custom.number_format = 'DD MMM YYYY'

    _save(wb, "p4_date_formats")


# ===================================================================
# Main
# ===================================================================

if __name__ == "__main__":
    print("Generating synthetic XLSX fixtures...")

    # P1: Tidy Analyst
    build_p1_simple_financial()
    build_p1_inventory_mixed()
    build_p1_time_series()
    build_p1_multi_sheet()

    # P2: Merger
    build_p2_budget_merged()
    build_p2_deep_hierarchy()
    build_p2_cross_tab()
    build_p2_irregular_merges()

    # P3: Multi-Tasker
    build_p3_two_tables_blank_rows()
    build_p3_tables_with_titles()
    build_p3_tables_with_footnotes()
    build_p3_side_by_side()

    # P4: Formatter
    build_p4_kpi_dashboard()
    build_p4_banded_report()
    build_p4_hidden_columns()
    build_p4_date_formats()

    count = len(list(OUT.glob("*.xlsx")))
    print(f"\nDone: {count} files in {OUT}")
