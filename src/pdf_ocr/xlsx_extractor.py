"""Excel (XLSX/XLS) table extraction.

Extracts structured tables from Excel files.
- XLSX (Office 2007+): Uses openpyxl
- XLS (Office 97-2003): Uses xlrd

XLSX provides explicit grid structure, so geometry heuristics are minimal.
Semantic heuristics (header detection, type patterns) from heuristics.py apply.

Visual elements (fills, borders, fonts) are extracted for cross-validation
with text-based heuristics, following the dual-channel validation philosophy.
Note: Visual extraction is limited for XLS files (xlrd has minimal style support).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

from pdf_ocr.heuristics import (
    CellType,
    StructuredTable,
    TableMetadata,
    build_column_names_from_headers,
    detect_cell_type,
    estimate_header_rows,
    is_header_type_pattern,
    normalize_grid,
    split_headers_from_data,
)

if TYPE_CHECKING:
    from openpyxl.cell import Cell
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Visual element structures (VH1-VH6 for XLSX)
# ---------------------------------------------------------------------------


@dataclass
class ExcelCellStyle:
    """Visual style information for a cell."""

    fill_color: tuple[float, float, float] | None = None  # RGB 0-1
    border_top: bool = False
    border_bottom: bool = False
    border_left: bool = False
    border_right: bool = False
    font_bold: bool = False
    font_italic: bool = False
    font_size: float | None = None
    font_color: tuple[float, float, float] | None = None


@dataclass
class ExcelVisualInfo:
    """Visual information for the entire table."""

    header_fill_rows: list[int] = field(default_factory=list)  # VH2: Header fills
    zebra_pattern: bool = False  # VH3: Alternating row colors
    has_grid_borders: bool = False  # VH1: Grid structure
    separator_rows: list[int] = field(default_factory=list)  # VH4: Section separators


@dataclass
class ExcelTable:
    """A table extracted from an Excel sheet with visual info."""

    grid: list[list[str]]
    styles: list[list[ExcelCellStyle]] | None = None
    visual: ExcelVisualInfo | None = None
    sheet_name: str = ""
    table_index: int = 0


# ---------------------------------------------------------------------------
# Color utilities
# ---------------------------------------------------------------------------


def _parse_color(color) -> tuple[float, float, float] | None:
    """Parse openpyxl color to RGB tuple (0-1 range).

    Handles various color formats: theme colors, indexed colors, RGB.
    Returns None for transparent/no-fill.
    """
    if color is None:
        return None

    # Check for RGB color
    if hasattr(color, "rgb") and color.rgb:
        rgb = color.rgb
        # Sometimes rgb is "00RRGGBB" format
        if isinstance(rgb, str) and len(rgb) >= 6:
            # Take last 6 characters (RRGGBB), skipping alpha if present
            hex_color = rgb[-6:]
            try:
                r = int(hex_color[0:2], 16) / 255.0
                g = int(hex_color[2:4], 16) / 255.0
                b = int(hex_color[4:6], 16) / 255.0
                # Skip pure white and transparent
                if (r, g, b) == (1.0, 1.0, 1.0):
                    return None
                return (r, g, b)
            except ValueError:
                return None

    # Check for indexed color (legacy)
    if hasattr(color, "indexed") and color.indexed is not None:
        # Indexed colors are tricky; for now skip them
        return None

    # Check for theme color (needs workbook for resolution)
    if hasattr(color, "theme") and color.theme is not None:
        # Theme colors require workbook context; skip for now
        return None

    return None


def _extract_cell_style(cell: "Cell") -> ExcelCellStyle:
    """Extract visual style information from a cell."""
    style = ExcelCellStyle()

    # Fill color (VH2/VH3/VH5)
    if cell.fill and cell.fill.patternType and cell.fill.patternType != "none":
        fg_color = getattr(cell.fill, "fgColor", None)
        style.fill_color = _parse_color(fg_color)

    # Borders (VH1)
    if cell.border:
        style.border_top = cell.border.top and cell.border.top.style is not None
        style.border_bottom = cell.border.bottom and cell.border.bottom.style is not None
        style.border_left = cell.border.left and cell.border.left.style is not None
        style.border_right = cell.border.right and cell.border.right.style is not None

    # Font (FH1-FH6)
    if cell.font:
        style.font_bold = bool(cell.font.bold)
        style.font_italic = bool(cell.font.italic)
        style.font_size = cell.font.size
        style.font_color = _parse_color(cell.font.color)

    return style


# ---------------------------------------------------------------------------
# Visual heuristics
# ---------------------------------------------------------------------------


def _analyze_visual_structure(
    styles: list[list[ExcelCellStyle]],
) -> ExcelVisualInfo:
    """Run visual heuristics (VH1-VH6) on extracted styles."""
    visual = ExcelVisualInfo()

    if not styles or not styles[0]:
        return visual

    num_rows = len(styles)

    # VH1: Grid detection - check if most cells have borders
    border_count = 0
    total_cells = 0
    for row in styles:
        for style in row:
            total_cells += 1
            if any([style.border_top, style.border_bottom,
                    style.border_left, style.border_right]):
                border_count += 1
    visual.has_grid_borders = total_cells > 0 and border_count / total_cells > 0.5

    # VH2: Header fill detection - first few rows with distinct fill color
    row_fills: list[tuple[float, float, float] | None] = []
    for ri, row in enumerate(styles):
        # Get most common fill in the row
        fills = [s.fill_color for s in row if s.fill_color is not None]
        if fills:
            # Use first non-null fill as representative
            row_fills.append(fills[0])
        else:
            row_fills.append(None)

    # Find header rows (first N rows with consistent fill, different from data)
    if row_fills:
        first_fill = row_fills[0]
        if first_fill is not None:
            for ri, fill in enumerate(row_fills[:5]):  # Check first 5 rows
                if fill == first_fill:
                    visual.header_fill_rows.append(ri)
                else:
                    break

    # VH3: Zebra striping detection
    if num_rows >= 6:
        # Skip potential header rows
        start_check = max(2, len(visual.header_fill_rows))
        fills_in_data = row_fills[start_check:]

        if len(fills_in_data) >= 4:
            # Check for alternating pattern
            unique_fills = set(fills_in_data)
            if len(unique_fills) == 2 and None not in unique_fills:
                # Check if actually alternating
                alternates = True
                for i in range(1, len(fills_in_data)):
                    if fills_in_data[i] == fills_in_data[i - 1]:
                        alternates = False
                        break
                visual.zebra_pattern = alternates

    # VH4: Section separators - rows with thick bottom borders
    for ri, row in enumerate(styles):
        thick_bottom_count = sum(
            1 for s in row if s.border_bottom
        )
        if thick_bottom_count > len(row) * 0.8:
            # Check if this is different from regular grid borders
            if not visual.has_grid_borders:
                visual.separator_rows.append(ri)

    return visual


# ---------------------------------------------------------------------------
# Merged cell handling
# ---------------------------------------------------------------------------


def _expand_merged_cells(
    ws: "Worksheet",
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> dict[tuple[int, int], str]:
    """Build a map of cell values accounting for merged regions.

    For merged cells, the value is stored in the top-left cell.
    This function returns a mapping from (row, col) to the displayed value.
    """
    cell_values: dict[tuple[int, int], str] = {}

    # First pass: get all regular cell values
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                cell_values[(row, col)] = str(value)
            else:
                cell_values[(row, col)] = ""

    # Second pass: handle merged regions
    for merged_range in ws.merged_cells.ranges:
        # Get the value from the top-left cell
        min_r, min_c, max_r, max_c = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )

        # Check if this merge is within our range
        if min_r > max_row or max_r < min_row or min_c > max_col or max_c < min_col:
            continue

        top_left_value = cell_values.get((min_r, min_c), "")

        # Fill all cells in the merge with the same value
        # (for header stacking, we only need the top-left)
        # Mark other cells as belonging to a merge
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r >= min_row and r <= max_row and c >= min_col and c <= max_col:
                    if (r, c) != (min_r, min_c):
                        # Mark as merged (empty, value is in top-left)
                        cell_values[(r, c)] = ""

    return cell_values


# ---------------------------------------------------------------------------
# Table extraction
# ---------------------------------------------------------------------------


def _find_data_bounds(ws: "Worksheet") -> tuple[int, int, int, int] | None:
    """Find the bounding box of actual data in the worksheet.

    Returns (min_row, max_row, min_col, max_col) or None if empty.
    """
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column

    if min_row is None or max_row is None or min_col is None or max_col is None:
        return None

    # Trim trailing empty rows/columns
    # (openpyxl sometimes reports dimensions including formatted but empty cells)
    # For simplicity, we trust the worksheet bounds for now
    return (min_row, max_row, min_col, max_col)


def _extract_table_from_range(
    ws: "Worksheet",
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    extract_styles: bool = True,
) -> ExcelTable:
    """Extract a table from a specific range in the worksheet."""
    # Get cell values accounting for merges
    cell_values = _expand_merged_cells(ws, min_row, max_row, min_col, max_col)

    # Build grid
    grid: list[list[str]] = []
    styles: list[list[ExcelCellStyle]] = []

    for row in range(min_row, max_row + 1):
        row_data: list[str] = []
        row_styles: list[ExcelCellStyle] = []

        for col in range(min_col, max_col + 1):
            row_data.append(cell_values.get((row, col), ""))

            if extract_styles:
                cell = ws.cell(row=row, column=col)
                row_styles.append(_extract_cell_style(cell))

        grid.append(row_data)
        if extract_styles:
            styles.append(row_styles)

    # Analyze visual structure
    visual = _analyze_visual_structure(styles) if styles else None

    return ExcelTable(
        grid=grid,
        styles=styles if extract_styles else None,
        visual=visual,
        sheet_name=ws.title or "",
    )


def _detect_table_regions(
    ws: "Worksheet",
    bounds: tuple[int, int, int, int],
) -> list[tuple[int, int, int, int]]:
    """Detect separate table regions within a worksheet.

    Looks for gaps (multiple empty rows/columns) that separate tables.
    Most worksheets have a single table, but some have multiple.

    Returns list of (min_row, max_row, min_col, max_col) tuples.
    """
    min_row, max_row, min_col, max_col = bounds

    # For now, treat entire range as one table
    # TODO: Implement gap detection for multiple tables
    return [(min_row, max_row, min_col, max_col)]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def extract_tables_from_xlsx(
    xlsx_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
    extract_styles: bool = True,
) -> list[StructuredTable]:
    """Extract structured tables from an Excel file.

    Args:
        xlsx_path: Path to the .xlsx file
        sheets: Optional list of sheet names or indices to process.
            If None, processes all sheets.
        extract_styles: If True, extract visual style information.

    Returns:
        List of StructuredTable objects, one per table found.

    Raises:
        ImportError: If openpyxl is not installed.
        FileNotFoundError: If the file doesn't exist.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX extraction. "
            "Install with: pip install pdf-ocr[xlsx]"
        ) from e

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    # Load workbook (read-only for performance)
    wb: Workbook = load_workbook(path, read_only=False, data_only=True)

    result: list[StructuredTable] = []

    # Determine which sheets to process
    if sheets is None:
        sheet_list = wb.sheetnames
    else:
        sheet_list = []
        for s in sheets:
            if isinstance(s, int):
                if 0 <= s < len(wb.sheetnames):
                    sheet_list.append(wb.sheetnames[s])
            else:
                if s in wb.sheetnames:
                    sheet_list.append(s)

    for sheet_idx, sheet_name in enumerate(sheet_list):
        ws: Worksheet = wb[sheet_name]

        # Find data bounds
        bounds = _find_data_bounds(ws)
        if bounds is None:
            continue  # Empty sheet

        # Detect table regions (usually just one)
        regions = _detect_table_regions(ws, bounds)

        for table_idx, (min_row, max_row, min_col, max_col) in enumerate(regions):
            # Extract raw table
            excel_table = _extract_table_from_range(
                ws, min_row, max_row, min_col, max_col,
                extract_styles=extract_styles,
            )
            excel_table.table_index = table_idx

            if not excel_table.grid or not excel_table.grid[0]:
                continue

            # Normalize grid
            grid = normalize_grid(excel_table.grid)

            # Estimate header rows
            # Use visual info if available, otherwise use heuristics
            header_count = 0
            if excel_table.visual and excel_table.visual.header_fill_rows:
                # Visual-based header detection
                header_count = len(excel_table.visual.header_fill_rows)
            else:
                # Heuristic-based header detection
                header_count = estimate_header_rows(grid)

            # Split headers from data
            header_rows, data_rows = split_headers_from_data(grid, header_count)

            # Build column names
            column_names = build_column_names_from_headers(header_rows)

            # Ensure column names list has correct length
            num_cols = len(grid[0]) if grid else 0
            while len(column_names) < num_cols:
                column_names.append("")

            # Create metadata
            metadata = TableMetadata(
                page_number=sheet_idx,  # Use sheet index as "page"
                table_index=table_idx,
                section_label=None,
                sheet_name=sheet_name,
            )

            # Create StructuredTable
            result.append(StructuredTable(
                column_names=column_names,
                data=data_rows,
                source_format="xlsx",
                metadata={
                    "page_number": metadata.page_number,
                    "table_index": metadata.table_index,
                    "sheet_name": metadata.sheet_name,
                },
            ))

    wb.close()
    return result


# ---------------------------------------------------------------------------
# XLS (Legacy Excel 97-2003) extraction using xlrd
# ---------------------------------------------------------------------------


def _extract_tables_from_xls(
    xls_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
) -> list[StructuredTable]:
    """Extract structured tables from a legacy XLS file.

    Args:
        xls_path: Path to the .xls file
        sheets: Optional list of sheet names or indices to process.

    Returns:
        List of StructuredTable objects.

    Note: Visual style extraction is limited for XLS files.
    """
    try:
        import xlrd
    except ImportError as e:
        raise ImportError(
            "xlrd is required for XLS extraction. "
            "Install with: pip install pdf-ocr[xls]"
        ) from e

    path = Path(xls_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {xls_path}")

    wb = xlrd.open_workbook(path, formatting_info=True)
    result: list[StructuredTable] = []

    # Determine which sheets to process
    if sheets is None:
        sheet_indices = range(wb.nsheets)
    else:
        sheet_indices = []
        for s in sheets:
            if isinstance(s, int):
                if 0 <= s < wb.nsheets:
                    sheet_indices.append(s)
            else:
                try:
                    idx = wb.sheet_names().index(s)
                    sheet_indices.append(idx)
                except ValueError:
                    pass

    for sheet_idx in sheet_indices:
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name

        if ws.nrows == 0 or ws.ncols == 0:
            continue

        # Build grid
        grid: list[list[str]] = []
        for row_idx in range(ws.nrows):
            row_data: list[str] = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                # Convert cell value to string
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row_data.append("")
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    # Convert Excel date to string
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        row_data.append(f"{date_tuple[0]:04d}-{date_tuple[1]:02d}-{date_tuple[2]:02d}")
                    except Exception:
                        row_data.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Format number nicely
                    if cell.value == int(cell.value):
                        row_data.append(str(int(cell.value)))
                    else:
                        row_data.append(str(cell.value))
                else:
                    row_data.append(str(cell.value).strip())
            grid.append(row_data)

        if not grid or not grid[0]:
            continue

        # Normalize and process
        grid = normalize_grid(grid)
        header_count = estimate_header_rows(grid)
        header_rows, data_rows = split_headers_from_data(grid, header_count)
        column_names = build_column_names_from_headers(header_rows)

        num_cols = len(grid[0]) if grid else 0
        while len(column_names) < num_cols:
            column_names.append("")

        result.append(StructuredTable(
            column_names=column_names,
            data=data_rows,
            source_format="xls",
            metadata={
                "page_number": sheet_idx,
                "table_index": 0,
                "sheet_name": sheet_name,
            },
        ))

    return result


def extract_tables_from_excel(
    excel_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
    extract_styles: bool = True,
) -> list[StructuredTable]:
    """Extract structured tables from an Excel file (XLS or XLSX).

    Auto-detects format based on file extension and uses the appropriate
    extraction method.

    Args:
        excel_path: Path to the Excel file (.xlsx or .xls)
        sheets: Optional list of sheet names or indices to process.
        extract_styles: If True, extract visual styles (XLSX only).

    Returns:
        List of StructuredTable objects.
    """
    path = Path(excel_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return extract_tables_from_xlsx(path, sheets=sheets, extract_styles=extract_styles)
    elif suffix == ".xls":
        return _extract_tables_from_xls(path, sheets=sheets)
    else:
        # Try XLSX first (more common), fall back to XLS
        try:
            return extract_tables_from_xlsx(path, sheets=sheets, extract_styles=extract_styles)
        except Exception:
            return _extract_tables_from_xls(path, sheets=sheets)


def xlsx_to_markdown(
    xlsx_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
) -> str:
    """Extract tables from Excel and render as markdown.

    Convenience function for quick inspection of Excel content.
    """
    tables = extract_tables_from_xlsx(xlsx_path, sheets=sheets, extract_styles=False)

    parts: list[str] = []
    for table in tables:
        meta = table.metadata or {}
        sheet = meta.get("sheet_name", "Unknown")
        parts.append(f"## Sheet: {sheet}\n")

        if table.column_names:
            parts.append("| " + " | ".join(table.column_names) + " |")
            parts.append("| " + " | ".join(["---"] * len(table.column_names)) + " |")

        for row in table.data:
            parts.append("| " + " | ".join(row) + " |")

        parts.append("")  # Blank line between tables

    return "\n".join(parts)
