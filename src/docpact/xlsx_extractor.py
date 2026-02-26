"""Excel (XLSX/XLS) table extraction.

Extracts structured tables from Excel files.
- XLSX (Office 2007+): Uses openpyxl
- XLS (Office 97-2003): Uses xlrd

XLSX provides explicit grid structure, so geometry heuristics are minimal.
Semantic heuristics (header detection, type patterns) from heuristics.py apply.

Visual elements (fills, borders, fonts) are extracted for cross-validation
with text-based heuristics, following the dual-channel validation philosophy.
Note: Visual extraction is limited for XLS files (xlrd has minimal style support).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

from docpact.heuristics import (
    CellType,
    StructuredTable,
    TableMetadata,
    build_column_names_from_headers,
    detect_cell_type,
    estimate_header_rows,
    estimate_header_rows_from_types,
    is_header_type_pattern,
    normalize_grid,
    split_headers_from_data,
)

if TYPE_CHECKING:
    from openpyxl.cell import Cell
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Visual element structures (VH1-VH6 for XLSX)
# ---------------------------------------------------------------------------


@dataclass
class ExcelCellStyle:
    """Visual style information for a cell."""

    fill_color: tuple[float, float, float] | None = None  # RGB 0-1
    border_top: bool = False
    border_bottom: bool = False
    border_left: bool = False
    border_right: bool = False
    font_bold: bool = False
    font_italic: bool = False
    font_size: float | None = None
    font_color: tuple[float, float, float] | None = None


@dataclass
class ExcelVisualInfo:
    """Visual information for the entire table."""

    header_fill_rows: list[int] = field(default_factory=list)  # VH2: Header fills
    zebra_pattern: bool = False  # VH3: Alternating row colors
    has_grid_borders: bool = False  # VH1: Grid structure
    separator_rows: list[int] = field(default_factory=list)  # VH4: Section separators


@dataclass
class ExcelTable:
    """A table extracted from an Excel sheet with visual info."""

    grid: list[list[str]]
    styles: list[list[ExcelCellStyle]] | None = None
    visual: ExcelVisualInfo | None = None
    sheet_name: str = ""
    table_index: int = 0
    number_format_hints: list[list[str | None]] = field(default_factory=list)  # XH4


# ---------------------------------------------------------------------------
# Color utilities
# ---------------------------------------------------------------------------


def _parse_color(color) -> tuple[float, float, float] | None:
    """Parse openpyxl color to RGB tuple (0-1 range).

    Handles various color formats: theme colors, indexed colors, RGB.
    Returns None for transparent/no-fill.
    """
    if color is None:
        return None

    # Check for RGB color
    if hasattr(color, "rgb") and color.rgb:
        rgb = color.rgb
        # Sometimes rgb is "00RRGGBB" format
        if isinstance(rgb, str) and len(rgb) >= 6:
            # Take last 6 characters (RRGGBB), skipping alpha if present
            hex_color = rgb[-6:]
            try:
                r = int(hex_color[0:2], 16) / 255.0
                g = int(hex_color[2:4], 16) / 255.0
                b = int(hex_color[4:6], 16) / 255.0
                # Skip pure white and transparent
                if (r, g, b) == (1.0, 1.0, 1.0):
                    return None
                return (r, g, b)
            except ValueError:
                return None

    # Check for indexed color (legacy)
    if hasattr(color, "indexed") and color.indexed is not None:
        # Indexed colors are tricky; for now skip them
        return None

    # Check for theme color (needs workbook for resolution)
    if hasattr(color, "theme") and color.theme is not None:
        # Theme colors require workbook context; skip for now
        return None

    return None


def _extract_cell_style(cell: "Cell") -> ExcelCellStyle:
    """Extract visual style information from a cell."""
    style = ExcelCellStyle()

    # Fill color (VH2/VH3/VH5)
    if cell.fill and cell.fill.patternType and cell.fill.patternType != "none":
        fg_color = getattr(cell.fill, "fgColor", None)
        style.fill_color = _parse_color(fg_color)

    # Borders (VH1)
    if cell.border:
        style.border_top = cell.border.top and cell.border.top.style is not None
        style.border_bottom = cell.border.bottom and cell.border.bottom.style is not None
        style.border_left = cell.border.left and cell.border.left.style is not None
        style.border_right = cell.border.right and cell.border.right.style is not None

    # Font (FH1-FH6)
    if cell.font:
        style.font_bold = bool(cell.font.bold)
        style.font_italic = bool(cell.font.italic)
        style.font_size = cell.font.size
        style.font_color = _parse_color(cell.font.color)

    return style


# ---------------------------------------------------------------------------
# Visual heuristics
# ---------------------------------------------------------------------------


def _analyze_visual_structure(
    styles: list[list[ExcelCellStyle]],
) -> ExcelVisualInfo:
    """Run visual heuristics (VH1-VH6) on extracted styles."""
    visual = ExcelVisualInfo()

    if not styles or not styles[0]:
        return visual

    num_rows = len(styles)

    # VH1: Grid detection - check if most cells have borders
    border_count = 0
    total_cells = 0
    for row in styles:
        for style in row:
            total_cells += 1
            if any([style.border_top, style.border_bottom,
                    style.border_left, style.border_right]):
                border_count += 1
    visual.has_grid_borders = total_cells > 0 and border_count / total_cells > 0.5

    # VH2: Header fill detection - first few rows with distinct fill color
    row_fills: list[tuple[float, float, float] | None] = []
    for ri, row in enumerate(styles):
        # Get most common fill in the row
        fills = [s.fill_color for s in row if s.fill_color is not None]
        if fills:
            # Use first non-null fill as representative
            row_fills.append(fills[0])
        else:
            row_fills.append(None)

    # Find header rows (first N rows with consistent fill, different from data)
    if row_fills:
        first_fill = row_fills[0]
        if first_fill is not None:
            for ri, fill in enumerate(row_fills[:5]):  # Check first 5 rows
                if fill == first_fill:
                    visual.header_fill_rows.append(ri)
                else:
                    break

    # VH3: Zebra striping detection
    if num_rows >= 6:
        # Skip potential header rows
        start_check = max(2, len(visual.header_fill_rows))
        fills_in_data = row_fills[start_check:]

        if len(fills_in_data) >= 4:
            # Check for alternating pattern
            unique_fills = set(fills_in_data)
            if len(unique_fills) == 2 and None not in unique_fills:
                # Check if actually alternating
                alternates = True
                for i in range(1, len(fills_in_data)):
                    if fills_in_data[i] == fills_in_data[i - 1]:
                        alternates = False
                        break
                visual.zebra_pattern = alternates

    # VH4: Section separators - rows with thick bottom borders
    for ri, row in enumerate(styles):
        thick_bottom_count = sum(
            1 for s in row if s.border_bottom
        )
        if thick_bottom_count > len(row) * 0.8:
            # Check if this is different from regular grid borders
            if not visual.has_grid_borders:
                visual.separator_rows.append(ri)

    return visual


# ---------------------------------------------------------------------------
# Merged cell handling
# ---------------------------------------------------------------------------


def _expand_merged_cells(
    ws: "Worksheet",
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> dict[tuple[int, int], str]:
    """Build a map of cell values accounting for merged regions.

    For merged cells, the value is stored in the top-left cell.
    This function returns a mapping from (row, col) to the displayed value.
    """
    cell_values: dict[tuple[int, int], str] = {}

    # First pass: get all regular cell values
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                cell_values[(row, col)] = str(value)
            else:
                cell_values[(row, col)] = ""

    # Second pass: handle merged regions
    for merged_range in ws.merged_cells.ranges:
        # Get the value from the top-left cell
        min_r, min_c, max_r, max_c = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )

        # Check if this merge is within our range
        if min_r > max_row or max_r < min_row or min_c > max_col or max_c < min_col:
            continue

        top_left_value = cell_values.get((min_r, min_c), "")

        # Fill all cells in the merge with the same value
        # (for header stacking, we only need the top-left)
        # Mark other cells as belonging to a merge
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r >= min_row and r <= max_row and c >= min_col and c <= max_col:
                    if (r, c) != (min_r, min_c):
                        # Mark as merged (empty, value is in top-left)
                        cell_values[(r, c)] = ""

    return cell_values


# ---------------------------------------------------------------------------
# Table extraction
# ---------------------------------------------------------------------------


def _find_data_bounds(
    ws: "Worksheet",
    *,
    filter_hidden: bool = True,
) -> tuple[int, int, int, int] | None:
    """Find the bounding box of actual data in the worksheet.

    Returns (min_row, max_row, min_col, max_col) or None if empty.
    When filter_hidden=True, hidden rows/columns are excluded from bounds.
    """
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column

    if min_row is None or max_row is None or min_col is None or max_col is None:
        return None

    # Trim trailing empty rows/columns
    # (openpyxl sometimes reports dimensions including formatted but empty cells)
    # For simplicity, we trust the worksheet bounds for now
    return (min_row, max_row, min_col, max_col)


def _get_hidden_cols(ws: "Worksheet", min_col: int, max_col: int) -> set[int]:
    """XH3: Get set of hidden column numbers.

    Checks ws.column_dimensions[letter].hidden for each column.
    """
    from openpyxl.utils import get_column_letter

    hidden: set[int] = set()
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.hidden:
            hidden.add(col)
    return hidden


def _get_hidden_rows(ws: "Worksheet", min_row: int, max_row: int) -> set[int]:
    """XH3: Get set of hidden row numbers.

    Checks ws.row_dimensions[row].hidden for each row.
    """
    hidden: set[int] = set()
    for row in range(min_row, max_row + 1):
        dim = ws.row_dimensions.get(row)
        if dim and dim.hidden:
            hidden.add(row)
    return hidden


def _detect_number_format_hint(cell: "Cell") -> str | None:
    """XH4: Infer a type hint from the cell's number format.

    Returns 'date', 'currency', 'percentage', or None.
    """
    fmt = cell.number_format
    if not fmt or fmt == "General":
        return None

    fmt_lower = fmt.lower()

    # Date patterns
    date_indicators = ["yyyy", "yy", "mm", "dd", "mmm", "mmmm"]
    if any(ind in fmt_lower for ind in date_indicators):
        # Exclude time-only formats
        if any(d in fmt_lower for d in ["y", "d"]):
            return "date"

    # Currency patterns
    if any(sym in fmt for sym in ["$", "€", "£", "¥"]):
        return "currency"

    # Percentage
    if "%" in fmt:
        return "percentage"

    return None


def _estimate_header_rows_from_merges(
    ws: "Worksheet",
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    grid: list[list[str]],
    *,
    max_scan: int = 10,
) -> int:
    """Estimate header row count using horizontal merge ranges + type continuation.

    Only counts merges that span multiple columns (horizontal spans) as header
    indicators. Vertical-only merges (spanning rows but not columns) indicate
    row label grouping, not headers.

    1. Find the last row (within first max_scan rows) with a horizontal merge.
    2. Continue past the last merge row while rows are all-string (TH2).
    3. Return the total count.

    This mirrors DOCX's DH2 but uses openpyxl's merged_cells.ranges.
    """
    scan_limit = min(min_row + max_scan - 1, max_row)

    # Find last row with a horizontal merge in the first max_scan rows
    last_merge_row = -1
    for merged_range in ws.merged_cells.ranges:
        mr_min_r = merged_range.min_row
        mr_max_r = merged_range.max_row
        mr_min_c = merged_range.min_col
        mr_max_c = merged_range.max_col

        # Only count horizontal merges (spanning multiple columns)
        is_horizontal = mr_max_c > mr_min_c

        # Check if merge overlaps with our region's header zone
        if (is_horizontal and
                mr_min_r >= min_row and mr_min_r <= scan_limit and
                mr_min_c <= max_col and mr_max_c >= min_col):
            row_offset = mr_min_r - min_row
            if row_offset > last_merge_row:
                last_merge_row = row_offset

    if last_merge_row < 0:
        return 0  # No horizontal merges found

    # From last_merge_row + 1, continue while rows are all-string
    header_count = last_merge_row + 1
    for i in range(header_count, len(grid)):
        if is_header_type_pattern(grid[i]):
            header_count = i + 1
        else:
            break

    return header_count


def _build_column_names_with_forward_fill(
    header_rows: list[list[str]],
) -> list[str]:
    """Build column names from multi-row headers with forward-fill.

    For each header row, empty cells inherit the preceding non-empty cell's
    value (forward-fill). This handles merged header spans where a label like
    "Revenue" spans cols 2-3 but only col 2 has the value.

    Then stacks rows vertically with " / " separator, deduplicating consecutive
    identical fragments.
    """
    if not header_rows:
        return []

    num_cols = max(len(row) for row in header_rows)

    # Forward-fill each header row independently
    filled_rows: list[list[str]] = []
    for row in header_rows:
        filled: list[str] = []
        last_val = ""
        for ci in range(num_cols):
            val = row[ci].strip() if ci < len(row) else ""
            if val:
                last_val = val
                filled.append(val)
            else:
                filled.append(last_val)
        filled_rows.append(filled)

    # Stack rows with " / " separator, dedup consecutive identical fragments
    column_names: list[str] = []
    for ci in range(num_cols):
        parts: list[str] = []
        for row in filled_rows:
            val = row[ci] if ci < len(row) else ""
            if val and (not parts or val != parts[-1]):
                parts.append(val)
        column_names.append(" / ".join(parts))

    return column_names


def _detect_title_row(
    grid: list[list[str]],
    header_count: int,
) -> tuple[str | None, list[list[str]], int]:
    """XH2: Detect if the first row is a title row.

    A title row has exactly one non-empty cell and header_count > 1.
    Returns (title, adjusted_grid, adjusted_header_count).
    """
    if not grid or header_count <= 1:
        return None, grid, header_count

    non_empty = [c for c in grid[0] if c.strip()]
    if len(non_empty) == 1:
        title = non_empty[0]
        return title, grid[1:], header_count - 1

    return None, grid, header_count


def _extract_table_from_range(
    ws: "Worksheet",
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    extract_styles: bool = True,
    *,
    hidden_rows: set[int] | None = None,
    hidden_cols: set[int] | None = None,
) -> ExcelTable:
    """Extract a table from a specific range in the worksheet.

    XH3: When hidden_rows/hidden_cols are provided, those rows/columns
    are excluded from the output grid entirely.
    """
    if hidden_rows is None:
        hidden_rows = set()
    if hidden_cols is None:
        hidden_cols = set()

    # Get cell values accounting for merges
    cell_values = _expand_merged_cells(ws, min_row, max_row, min_col, max_col)

    # Build visible column list (XH3)
    visible_cols = [c for c in range(min_col, max_col + 1) if c not in hidden_cols]

    # Build grid
    grid: list[list[str]] = []
    styles: list[list[ExcelCellStyle]] = []
    number_format_hints: list[list[str | None]] = []

    for row in range(min_row, max_row + 1):
        if row in hidden_rows:
            continue  # XH3: skip hidden rows

        row_data: list[str] = []
        row_styles: list[ExcelCellStyle] = []
        row_hints: list[str | None] = []

        for col in visible_cols:
            row_data.append(cell_values.get((row, col), ""))

            cell = ws.cell(row=row, column=col)
            if extract_styles:
                row_styles.append(_extract_cell_style(cell))

            # XH4: number format hints
            row_hints.append(_detect_number_format_hint(cell))

        grid.append(row_data)
        if extract_styles:
            styles.append(row_styles)
        number_format_hints.append(row_hints)

    # Analyze visual structure
    visual = _analyze_visual_structure(styles) if styles else None

    return ExcelTable(
        grid=grid,
        styles=styles if extract_styles else None,
        visual=visual,
        sheet_name=ws.title or "",
        number_format_hints=number_format_hints,
    )


def _is_row_blank(
    ws: "Worksheet",
    row: int,
    min_col: int,
    max_col: int,
) -> bool:
    """Check if every cell in a row is empty or whitespace-only."""
    for col in range(min_col, max_col + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None and str(val).strip():
            return False
    return True


def _is_col_blank(
    ws: "Worksheet",
    col: int,
    min_row: int,
    max_row: int,
) -> bool:
    """Check if every cell in a column is empty or whitespace-only."""
    for row in range(min_row, max_row + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None and str(val).strip():
            return False
    return True


def _detect_table_regions(
    ws: "Worksheet",
    bounds: tuple[int, int, int, int],
    *,
    min_blank_rows: int = 2,
    min_blank_cols: int = 2,
    min_table_size: tuple[int, int] = (2, 2),
) -> list[tuple[int, int, int, int]]:
    """XH1: Detect separate table regions within a worksheet.

    Uses blank-row and blank-column gaps to separate tables:
    - Runs of N >= min_blank_rows consecutive blank rows → vertical separator
    - Within each vertical segment, runs of N >= min_blank_cols blank columns
      → horizontal separator

    Rationale: One blank row is breathing room within a table (subtotals,
    section breaks). Two+ blank rows signal "these are separate things."
    Same principle applies to columns.

    Returns list of (min_row, max_row, min_col, max_col) tuples.
    Each region has at least min_table_size cells.
    """
    min_row, max_row, min_col, max_col = bounds

    # Step 1: Find vertical segments (split by blank-row runs)
    row_segments = _split_by_blank_runs(
        [_is_row_blank(ws, r, min_col, max_col) for r in range(min_row, max_row + 1)],
        min_blank_rows,
    )
    # Convert relative offsets to absolute row numbers
    v_segments = [
        (min_row + start, min_row + end)
        for start, end in row_segments
    ]

    # Step 2: Within each vertical segment, find horizontal segments
    regions: list[tuple[int, int, int, int]] = []
    for seg_min_row, seg_max_row in v_segments:
        col_segments = _split_by_blank_runs(
            [_is_col_blank(ws, c, seg_min_row, seg_max_row) for c in range(min_col, max_col + 1)],
            min_blank_cols,
        )
        for col_start, col_end in col_segments:
            abs_min_col = min_col + col_start
            abs_max_col = min_col + col_end

            # Filter: region must be at least min_table_size
            num_rows = seg_max_row - seg_min_row + 1
            num_cols = abs_max_col - abs_min_col + 1
            if num_rows >= min_table_size[0] and num_cols >= min_table_size[1]:
                regions.append((seg_min_row, seg_max_row, abs_min_col, abs_max_col))

    if not regions:
        # Fallback: entire range as one table
        return [(min_row, max_row, min_col, max_col)]

    return regions


def _split_by_blank_runs(
    is_blank: list[bool],
    min_gap: int,
) -> list[tuple[int, int]]:
    """Split a sequence of blank/non-blank flags into non-blank segments.

    A run of >= min_gap consecutive True values is a separator.
    Returns list of (start_index, end_index) inclusive tuples for
    the non-blank segments.
    """
    n = len(is_blank)
    if n == 0:
        return []

    # Find runs of blanks
    segments: list[tuple[int, int]] = []
    seg_start: int | None = None

    for i in range(n):
        if not is_blank[i]:
            if seg_start is None:
                seg_start = i
        else:
            # Check if this is the start of a gap
            if seg_start is not None:
                # Count consecutive blanks from here
                gap_len = 0
                j = i
                while j < n and is_blank[j]:
                    gap_len += 1
                    j += 1
                if gap_len >= min_gap:
                    # End current segment
                    segments.append((seg_start, i - 1))
                    seg_start = None
                # If gap_len < min_gap, keep going (blank is within the table)

    # Close final segment
    if seg_start is not None:
        # Find last non-blank
        last_non_blank = seg_start
        for i in range(n - 1, seg_start - 1, -1):
            if not is_blank[i]:
                last_non_blank = i
                break
        segments.append((seg_start, last_non_blank))

    return segments


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def extract_tables_from_xlsx(
    xlsx_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
    extract_styles: bool = True,
    filter_hidden: bool = True,
) -> list[StructuredTable]:
    """Extract structured tables from an Excel file.

    Args:
        xlsx_path: Path to the .xlsx file
        sheets: Optional list of sheet names or indices to process.
            If None, processes all sheets.
        extract_styles: If True, extract visual style information.
        filter_hidden: If True (default), exclude hidden rows and columns.
            XH3 heuristic: content that the author chose not to show.

    Returns:
        List of StructuredTable objects, one per table found.

    Raises:
        ImportError: If openpyxl is not installed.
        FileNotFoundError: If the file doesn't exist.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX extraction. "
            "Install with: pip install docpact[xlsx]"
        ) from e

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    # Load workbook (read-only for performance)
    wb: Workbook = load_workbook(path, read_only=False, data_only=True)

    result: list[StructuredTable] = []

    # Determine which sheets to process
    if sheets is None:
        sheet_list = wb.sheetnames
    else:
        sheet_list = []
        for s in sheets:
            if isinstance(s, int):
                if 0 <= s < len(wb.sheetnames):
                    sheet_list.append(wb.sheetnames[s])
            else:
                if s in wb.sheetnames:
                    sheet_list.append(s)

    for sheet_idx, sheet_name in enumerate(sheet_list):
        ws: Worksheet = wb[sheet_name]

        # Find data bounds
        bounds = _find_data_bounds(ws)
        if bounds is None:
            continue  # Empty sheet

        # XH3: Identify hidden rows/columns
        hidden_rows: set[int] = set()
        hidden_cols: set[int] = set()
        if filter_hidden:
            hidden_rows = _get_hidden_rows(ws, bounds[0], bounds[1])
            hidden_cols = _get_hidden_cols(ws, bounds[2], bounds[3])

        # XH1: Detect table regions (handles multi-table sheets)
        regions = _detect_table_regions(ws, bounds)

        for table_idx, (min_row, max_row, min_col, max_col) in enumerate(regions):
            # Extract raw table (with XH3 hidden filtering)
            excel_table = _extract_table_from_range(
                ws, min_row, max_row, min_col, max_col,
                extract_styles=extract_styles,
                hidden_rows=hidden_rows,
                hidden_cols=hidden_cols,
            )
            excel_table.table_index = table_idx

            if not excel_table.grid or not excel_table.grid[0]:
                continue

            # Normalize grid
            grid = normalize_grid(excel_table.grid)

            # Estimate header rows — layered approach:
            # 1. Visual fill (if available and clear)
            # 2. Merge-based detection (last merge row + type continuation)
            # 3. Type-pattern analysis (TH2) — consecutive all-string rows
            # 4. Span-count analysis (H7) as last resort
            # Take the maximum across all methods.
            header_count = 0
            if excel_table.visual and excel_table.visual.header_fill_rows:
                header_count = len(excel_table.visual.header_fill_rows)
            else:
                # Merge-based: last merge row + continue while all-string
                merge_count = _estimate_header_rows_from_merges(
                    ws, min_row, max_row, min_col, max_col, grid,
                )
                # Type-pattern: consecutive all-string rows from top
                type_count = estimate_header_rows_from_types(grid)
                # Span-count: bottom-up pattern detection
                span_count = estimate_header_rows(grid)
                # Take the maximum — each method catches cases others miss
                header_count = max(merge_count, type_count, span_count)

            # XH2: Title row detection
            title, grid, header_count = _detect_title_row(grid, header_count)

            # Split headers from data
            header_rows, data_rows = split_headers_from_data(grid, header_count)

            # Build column names with forward-fill for merged header spans
            column_names = _build_column_names_with_forward_fill(header_rows)

            # Ensure column names list has correct length
            num_cols = len(grid[0]) if grid else 0
            while len(column_names) < num_cols:
                column_names.append("")

            # XH4: Collect number format hints for the table
            format_hints: dict[int, str] = {}
            if excel_table.number_format_hints:
                # Check first data row for format hints per column
                hint_start = header_count + (1 if title else 0)
                if hint_start < len(excel_table.number_format_hints):
                    for ci, hint in enumerate(excel_table.number_format_hints[hint_start]):
                        if hint:
                            format_hints[ci] = hint

            # Create metadata
            meta_dict: dict[str, str | int | None] = {
                "page_number": sheet_idx,
                "table_index": table_idx,
                "sheet_name": sheet_name,
            }
            if title:
                meta_dict["title"] = title
            if format_hints:
                meta_dict["format_hints"] = format_hints  # type: ignore[assignment]

            # Create StructuredTable
            result.append(StructuredTable(
                column_names=column_names,
                data=data_rows,
                source_format="xlsx",
                metadata=meta_dict,
            ))

    wb.close()
    return result


# ---------------------------------------------------------------------------
# XLS (Legacy Excel 97-2003) extraction using xlrd
# ---------------------------------------------------------------------------


def _extract_tables_from_xls(
    xls_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
) -> list[StructuredTable]:
    """Extract structured tables from a legacy XLS file.

    Args:
        xls_path: Path to the .xls file
        sheets: Optional list of sheet names or indices to process.

    Returns:
        List of StructuredTable objects.

    Note: Visual style extraction is limited for XLS files.
    """
    try:
        import xlrd
    except ImportError as e:
        raise ImportError(
            "xlrd is required for XLS extraction. "
            "Install with: pip install docpact[xls]"
        ) from e

    path = Path(xls_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {xls_path}")

    wb = xlrd.open_workbook(path, formatting_info=True)
    result: list[StructuredTable] = []

    # Determine which sheets to process
    if sheets is None:
        sheet_indices = range(wb.nsheets)
    else:
        sheet_indices = []
        for s in sheets:
            if isinstance(s, int):
                if 0 <= s < wb.nsheets:
                    sheet_indices.append(s)
            else:
                try:
                    idx = wb.sheet_names().index(s)
                    sheet_indices.append(idx)
                except ValueError:
                    pass

    for sheet_idx in sheet_indices:
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name

        if ws.nrows == 0 or ws.ncols == 0:
            continue

        # Build grid
        grid: list[list[str]] = []
        for row_idx in range(ws.nrows):
            row_data: list[str] = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                # Convert cell value to string
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row_data.append("")
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    # Convert Excel date to string
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        row_data.append(f"{date_tuple[0]:04d}-{date_tuple[1]:02d}-{date_tuple[2]:02d}")
                    except Exception:
                        row_data.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Format number nicely
                    if cell.value == int(cell.value):
                        row_data.append(str(int(cell.value)))
                    else:
                        row_data.append(str(cell.value))
                else:
                    row_data.append(str(cell.value).strip())
            grid.append(row_data)

        if not grid or not grid[0]:
            continue

        # Normalize and process
        grid = normalize_grid(grid)
        header_count = estimate_header_rows(grid)
        header_rows, data_rows = split_headers_from_data(grid, header_count)
        column_names = build_column_names_from_headers(header_rows)

        num_cols = len(grid[0]) if grid else 0
        while len(column_names) < num_cols:
            column_names.append("")

        result.append(StructuredTable(
            column_names=column_names,
            data=data_rows,
            source_format="xls",
            metadata={
                "page_number": sheet_idx,
                "table_index": 0,
                "sheet_name": sheet_name,
            },
        ))

    return result


def extract_tables_from_excel(
    excel_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
    extract_styles: bool = True,
) -> list[StructuredTable]:
    """Extract structured tables from an Excel file (XLS or XLSX).

    Auto-detects format based on file extension and uses the appropriate
    extraction method.

    Args:
        excel_path: Path to the Excel file (.xlsx or .xls)
        sheets: Optional list of sheet names or indices to process.
        extract_styles: If True, extract visual styles (XLSX only).

    Returns:
        List of StructuredTable objects.
    """
    path = Path(excel_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return extract_tables_from_xlsx(path, sheets=sheets, extract_styles=extract_styles)
    elif suffix == ".xls":
        return _extract_tables_from_xls(path, sheets=sheets)
    else:
        # Try XLSX first (more common), fall back to XLS
        try:
            return extract_tables_from_xlsx(path, sheets=sheets, extract_styles=extract_styles)
        except Exception:
            return _extract_tables_from_xls(path, sheets=sheets)


def xlsx_to_markdown(
    xlsx_path: str | Path,
    *,
    sheets: list[str | int] | None = None,
) -> str:
    """Extract tables from Excel and render as markdown.

    Convenience function for quick inspection of Excel content.
    """
    tables = extract_tables_from_xlsx(xlsx_path, sheets=sheets, extract_styles=False)

    parts: list[str] = []
    for table in tables:
        meta = table.metadata or {}
        sheet = meta.get("sheet_name", "Unknown")
        parts.append(f"## Sheet: {sheet}\n")

        if table.column_names:
            parts.append("| " + " | ".join(table.column_names) + " |")
            parts.append("| " + " | ".join(["---"] * len(table.column_names)) + " |")

        for row in table.data:
            parts.append("| " + " | ".join(row) + " |")

        parts.append("")  # Blank line between tables

    return "\n".join(parts)
